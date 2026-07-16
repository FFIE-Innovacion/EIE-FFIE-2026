#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
preparar_datos.py — Regenera la capa de datos pública del Dashboard IEIE v3.
Uso:
  python preparar_datos.py --base FFIE_2026_Base_Unida_IEIE.xlsx \
                           --marco "Poblacion_Total (13).xlsx" --out data/
Reglas aplicadas (ver docs/DATA_DICTIONARY.md y VALIDATION_REPORT.md):
- Excluir registros con Dato_Atipico=True (atipicos criticos).
- 999 = "No calculable": se excluye de todo promedio.
- El IEIE NO se recalcula; se promedian los valores oficiales.
- enc_departamento llega como codigo DANE numerico -> se mapea al nombre del marco.
- Suficiencia: muestra_valida >= 30. Cobertura = muestra/marco (acotada a 100% con nota).
- No se exporta ningun dato personal ni registro individual.
"""
import argparse, json, os
from collections import defaultdict, Counter
import openpyxl

UMBRAL_SUF = 30
DIMS = [("D1","D1_servicios_publicos_basicos"),("D2","D2_antiguedad_estado_general"),
("D3","D3_accesibilidad_entorno_seguridad"),("D4","D4_ambientes_aprendizaje_capacidad"),
("D5","D5_ambientes_inhabilitados_deficiencias"),("D6","D6_condiciones_fisicas_deficiencias"),
("D7","D7_confort_fisico"),("D8","D8_dotacion_mobiliario"),("D9","D9_afectacion_academica")]
DIM_LBL = {"D1":"Servicios públicos básicos","D2":"Antigüedad y estado general",
"D3":"Accesibilidad, entorno y seguridad","D4":"Ambientes de aprendizaje y capacidad",
"D5":"Ambientes inhabilitados","D6":"Condiciones físicas","D7":"Confort físico",
"D8":"Dotación y mobiliario","D9":"Afectación académica"}

def prom(a):
    a=[x for x in a if x is not None]
    return round(sum(a)/len(a),2) if a else None
def std(a):
    a=[x for x in a if x is not None]
    if len(a)<2: return None
    m=sum(a)/len(a); return round((sum((x-m)**2 for x in a)/(len(a)-1))**0.5,2)
def num(v):
    try:
        f=float(v); return None if f==999 else f
    except: return None
def cod2_of(v):
    v=str(v).strip()
    if v.endswith(".0"): v=v[:-2]
    return v.zfill(2)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--base",required=True); ap.add_argument("--marco",required=True)
    ap.add_argument("--out",default="data/")
    a=ap.parse_args(); os.makedirs(a.out,exist_ok=True)

    # ----- marco -----
    wb=openpyxl.load_workbook(a.marco,read_only=True); ws=wb["Total"]
    it=ws.iter_rows(values_only=True); hdr=[str(h).strip() if h else h for h in next(it)]
    ix={h:i for i,h in enumerate(hdr)}
    cod2_nombre={}; marco_dep=Counter(); marco_mpio=Counter(); marco_total=0
    for row in it:
        if row[ix["Código DANE Sede"]] is None: continue
        marco_total+=1
        mc=str(row[ix["Código municipio"]]).strip()
        if mc.endswith(".0"): mc=mc[:-2]
        mc=mc.zfill(5)
        cod2_nombre[mc[:2]]=str(row[ix["departamento"]]).strip()
        marco_dep[mc[:2]]+=1; marco_mpio[mc]+=1
    wb.close()
    def nombre_dep(c): return "Bogotá D.C." if c=="11" else cod2_nombre.get(c,c)

    # ----- base -----
    wb=openpyxl.load_workbook(a.base,read_only=True); ws=wb["ieie_respuestas"]
    it=ws.iter_rows(values_only=True); hdr=list(next(it)); ix={h:i for i,h in enumerate(hdr)}
    dep=defaultdict(lambda:{"ieie":[],"urb":[],"rur":[],"dims":defaultdict(list),
        "crit":Counter(),"fuerte":Counter(),"cat":Counter()})
    mpio=defaultdict(lambda:{"ieie":[],"cod2":None})
    nac={"ieie":[],"urb":[],"rur":[],"dims":defaultdict(list),"cat":Counter()}
    n999=Counter(); n_atip=0
    for row in it:
        if row[ix["Dato_Atipico"]] is True: n_atip+=1; continue
        c2=cod2_of(row[ix["enc_departamento"]])
        mc=str(row[ix["enc_municipio"]]).strip()
        if mc.endswith(".0"): mc=mc[:-2]
        mc=mc.zfill(5)
        D=dep[c2]; ieie=num(row[ix["IEIE_total"]])
        zona=str(row[ix["enc_tipo_de_predio"]]).strip().upper()
        if ieie is not None:
            D["ieie"].append(ieie); nac["ieie"].append(ieie)
            mpio[mc]["ieie"].append(ieie); mpio[mc]["cod2"]=c2
            if zona=="URBANO": D["urb"].append(ieie); nac["urb"].append(ieie)
            elif zona=="RURAL": D["rur"].append(ieie); nac["rur"].append(ieie)
        cat=str(row[ix["categoria_ieie"]]).strip()
        D["cat"][cat]+=1; nac["cat"][cat]+=1
        D["crit"][str(row[ix["dimension_mas_critica"]]).strip()]+=1
        D["fuerte"][str(row[ix["dimension_mas_fuerte"]]).strip()]+=1
        for cod,col in DIMS:
            v=num(row[ix[col]])
            if str(row[ix[col]]).strip()=="999": n999[cod]+=1
            D["dims"][cod].append(v); nac["dims"][cod].append(v)
    wb.close()

    n_val=len(nac["ieie"])
    resumen_nacional={"marco_total_sedes":marco_total,"muestra_valida":n_val,
      "atipicos_excluidos":n_atip,"cobertura_pct":round(100*n_val/marco_total,2),
      "ieie_nacional":prom(nac["ieie"]),"ieie_desv_std":std(nac["ieie"]),
      "ieie_urbano":prom(nac["urb"]),"ieie_rural":prom(nac["rur"]),
      "n_departamentos_marco":len(marco_dep),"n_departamentos_con_info":len(dep),
      "n_municipios_marco":len(marco_mpio),
      "n_municipios_con_info":len([m for m in mpio if mpio[m]["ieie"]]),
      "cobertura_territorial":"32 departamentos + Bogotá D.C.",
      "dist_categoria":dict(nac["cat"]),
      "dim_prom":{c:prom(nac["dims"][c]) for c,_ in DIMS}}
    deps=[]
    for c2,Dd in dep.items():
        nv=len(Dd["ieie"]); marco_n=marco_dep.get(c2,0); ieie=prom(Dd["ieie"])
        suf="suficiente" if nv>=UMBRAL_SUF else "insuficiente"
        pr = "no_evaluable" if suf=="insuficiente" else ("sin_info" if nv==0 else
             ("alta" if ieie<60 else ("media" if ieie<70 else "baja")))
        deps.append({"nombre":nombre_dep(c2),"cod":c2,"es_distrito_capital":c2=="11",
          "marco_sedes":marco_n,"muestra_valida":nv,
          "cobertura_pct":round(100*nv/marco_n,2) if marco_n else None,
          "ieie":ieie,"ieie_desv_std":std(Dd["ieie"]),
          "ieie_urbano":prom(Dd["urb"]),"ieie_rural":prom(Dd["rur"]),
          "suficiencia":suf,"prioridad_exploratoria":pr,
          "categoria_modal":Dd["cat"].most_common(1)[0][0] if Dd["cat"] else None,
          "dim_critica":Dd["crit"].most_common(1)[0][0] if Dd["crit"] else None,
          "dim_fuerte":Dd["fuerte"].most_common(1)[0][0] if Dd["fuerte"] else None,
          "dim_prom":{c:prom(Dd["dims"][c]) for c,_ in DIMS},
          "dist_categoria":dict(Dd["cat"])})
    deps.sort(key=lambda x:-(x["ieie"] or 0))
    munis=[]
    for mc,M in mpio.items():
        nv=len(M["ieie"]); marco_n=marco_mpio.get(mc,0)
        cob=round(100*nv/marco_n,2) if marco_n else None
        row={"mpio_cod":mc,"departamento":nombre_dep(M["cod2"]) if M["cod2"] else None,
          "marco_sedes":marco_n,"muestra_valida":nv,"cobertura_pct":cob,
          "ieie":prom(M["ieie"]),
          "suficiencia":"suficiente" if nv>=UMBRAL_SUF else "insuficiente"}
        if cob is not None and cob>100:
            row["cobertura_pct"]=100.0
            row["nota_cobertura"]="Muestra > marco por diferencias del código municipal; cobertura acotada a 100%."
        munis.append(row)
    def wr(n,o): json.dump(o,open(os.path.join(a.out,n),"w",encoding="utf-8"),
                           ensure_ascii=False,separators=(",",":"))
    wr("resumen_nacional.json",resumen_nacional)
    wr("resumen_departamental.json",deps)
    wr("resumen_municipal.json",munis)
    wr("perfil_dimensiones.json",{"no_calculable":999,
      "dimensiones":[{"codigo":c,"etiqueta":DIM_LBL[c],"prom_nacional":prom(nac["dims"][c]),
        "n_no_calculable":n999.get(c,0)} for c,_ in DIMS]})
    wr("resultados_rural_urbano.json",{"nacional":{"urbano":prom(nac["urb"]),
      "rural":prom(nac["rur"]),"n_urbano":len(nac["urb"]),"n_rural":len(nac["rur"])},
      "departamental":[{"nombre":d["nombre"],"urbano":d["ieie_urbano"],"rural":d["ieie_rural"]} for d in deps]})
    wr("suficiencia_muestral.json",{"umbral":UMBRAL_SUF,
      "criterio":"Muestra válida >= %d sedes por territorio"%UMBRAL_SUF,
      "departamentos":[{"nombre":d["nombre"],"cod":d["cod"],
        "muestra_valida":d["muestra_valida"],"suficiencia":d["suficiencia"]} for d in deps],
      "territorios_advertencia":[d["nombre"] for d in deps if d["suficiencia"]=="insuficiente"]})
    print("OK — marco:",marco_total,"| válidas:",n_val,"| atípicos:",n_atip,
          "| IEIE:",resumen_nacional["ieie_nacional"])
    print("NOTA: variables_detalle.json y módulos temáticos requieren el paso extendido",
          "(ver docs/DATA_DICTIONARY.md); este script cubre los agregados nucleares.")

if __name__=="__main__":
    main()
